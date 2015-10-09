#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Author: Alexander David Leech
Date:   07/10/2015
Rev:    0.3
Lang:   Python 2.7
Deps:   Pyserial, Pymodbus, scipy
"""

from pymodbus.client.sync import ModbusSerialClient as ModbusClient            #Import the MODBUS Protocol
from scipy.optimize import optimize as opt                                     #For the least squares algorithm
import logging                                                                 #For debugging purposes
import time                                                                    #Time based functions
import openpyxl as xl                                                          #For data logging in Excel



class xlsLogging:
    
    headers = ['PV', ' ', 'SP', 'OP']
    
    def __init__(self):
        #Load workbook and create new log sheet
        self.wb = xl.load_workbook('dataLogging.xlsx')
        self.ws = self.wb.create_sheet(title='Log' + str(len(self.wb.worksheets)+1))
       
        #Setup new sheet headers
        self.ws.cell(row = 0, column = 0).value = "Date: "
        self.ws.cell(row = 0, column = 1).value = time.ctime()
        self.ws.cell(row = 2, column = 0).value = "I" 
        for x in range(0, 4):
            self.ws.cell(row= 2, column= x+1).value = self.headers[x]
		
        #Initialise Iteration Number & EPOCH time
        self.i = 1
        self.startTime = time.time()
    
    
    def writeXls(self, data):
        
        self.ws.cell(row = self.i+2, column = 0).value = (time.time() - self.startTime)
        for x in range(0, 4):
            self.ws.cell(row = self.i+2, column = x+1).value = data.getRegister(x)
			print data.getRegister(x).rjust(x+1)
		
        self.wb.save('dataLogging.xlsx')
        self.i += 1        
        return 1
        
        
        
class comClient:

    interval = 5

    comSettings = {
                    "method"   : 'rtu',
                    "port"     : 'COM3',    #COM2 for virtual serial port tests & COM3/4 for online running
                    "stopbits" : 1,                
                    "bytesize" : 8,                
                    "parity"   : 'N',
                    "baudrate" : 9600,
                    "timeout"  : 1
                  }

    def __init__(self):
        self.datLog = xlsLogging()
        # Configure logging on the client 
        logging.basicConfig()
        self.log = logging.getLogger()
        #Change to DEBUG for full information during runtime
        self.log.setLevel(logging.INFO)
        #  Setup and Open the connection
        self.client = ModbusClient(**self.comSettings)

    def __getData(self):
        self.client.connect()
		rr = self.client.read_holding_registers(0, 4, unit=0x01)               #REMEMBER: Controller is unit 0x01!!!!!
        self.client.close()
        return rr

    def run(self):
        while(True):
            startTime = time.time()
			try:
		        self.datLog.writeXls(self.__getData())
			except:
				print "Modbus Error: Connection Failed"
				print "Attempting Reconnection..."
            
			time.sleep(self.interval - (time.time() - startTime))
        

		
if __name__ == '__main__':
	cc = comClient()
	
    cc.run()