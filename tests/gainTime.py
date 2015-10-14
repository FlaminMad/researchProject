#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Author: Alexander David Leech
Date:   12/10/2015
Rev:    1
Lang:   Python 2.7
Deps:   Pyserial, Pymodbus, scipy, openpyxl, collections
"""

from pymodbus.client.sync import ModbusSerialClient as ModbusClient            #Import the MODBUS Protocol
import logging                                                                 #For debugging purposes
import time                                                                    #Time based functions
import openpyxl as xl                                                          #For data logging in Excel


#-------------------------------------------------------------------------------
# Put the below in a different file.... when I can work out how to import it!!!!

from scipy import optimize as opt
import numpy

class leastSquares:
    
    #Initial guess of alpha and beta    
    x0 = numpy.array([1.0,1.0])
     
    def func(self,params,y,u,Y):
        #structure of function to be minimised
        return (Y - (params[0]*y + params[1]*u))


    def solve(self,y,u,Y):
        #run least squares algorithm - add try statement here
        ans = opt.leastsq(self.func,self.x0, args=(y,u,Y))
        #translate answers
        self.x0[0] = ans[0][0]
        self.x0[1] = ans[0][1]
        #for debugging - REMOVE LATER
        print self.x0

class xlsLogging:
    
    headers = ['PV', ' ', 'SP', 'OP']
    
    def __init__(self):
        #Load workbook and create new log sheet
        self.wb = xl.load_workbook('dataLogging.xlsx')
        self.ws = self.wb.create_sheet(title='Log' + str(len(self.wb.worksheets)+1))
       
        #Setup new sheet headers
        self.ws.cell(row = 0, column = 0).value = "Date: "
        self.ws.cell(row = 0, column = 1).value = time.ctime()
        self.ws.cell(row = 2, column = 0).value = "I" 
        for x in range(0, 4):
            self.ws.cell(row= 2, column= x+1).value = self.headers[x]
		
        #Initialise Iteration Number & EPOCH time
        self.i = 1
        self.startTime = time.time()
    
    
    def writeXls(self, data):
        
        self.ws.cell(row = self.i+2, column = 0).value = (time.time() - self.startTime)
        for x in range(0, 4):
            self.ws.cell(row = self.i+2, column = x+1).value = data.getRegister(x)
		
        self.wb.save('dataLogging.xlsx')
        self.i += 1        
        return 1
                
        
class comClient:

#Note controller parameters are listed here

    interval = 5
    sp = 500
    

    comSettings = {
                    "method"   : 'rtu',
                    "port"     : 'COM3',    #COM2 for virtual serial port tests & COM3/4 for online running
                    "stopbits" : 1,                
                    "bytesize" : 8,                
                    "parity"   : 'N',
                    "baudrate" : 9600,
                    "timeout"  : 1
                  }

    def __init__(self):
        # Configure logging on the client 
        logging.basicConfig()
        self.log = logging.getLogger()
        #Change to DEBUG for full information during runtime
        self.log.setLevel(logging.INFO)
        #  Setup and Open the connection
        self.client = ModbusClient(**self.comSettings)

    def getData(self):
        self.client.connect()
        rr = self.client.read_holding_registers(0, 4, unit=0x01)               #REMEMBER: Controller is unit 0x01!!!!!
        self.client.close()
        return rr

    def writeData(self,op):
        #Set to write data to the controller output (MODBUS address 3)
        self.client.connect()
        w = self.client.write_register(3,op,unit=0x01)
        self.client.close()
        return w
           
    def controller(self,a,b,y):
        ut = ((self.sp)-(a*y.getRegister(0)))/b
        print ut
        print y.getRegister(0)
        print ""
        if ut > 1000:
            self.writeData(1000)
        elif ut < 0:
            self.writeData(0)
        else:
            self.writeData(ut)
        

#-------------------------------------------------------------------------------
		
if __name__ == '__main__':
    cc = comClient()
    ls = leastSquares()
#    Uncomment to enable excel logging
#    datLog = xlsLogging()

    while(True):
#       For controller time keeping
        startTime = time.time()
        
#       Read controller data        
        try:
            rr = cc.getData()        
        except:
		print "Modbus Error: Connection Failed"
		break

#        Uncomment to enable excel logging of inputs
#        datLog.writeXls(r)

#       Interpret data
        ds.memoryUpdate(rr)

#       Update Outputs
        if ds.startupFlag == 0:
            ds.startupFlag = 1            
            continue
        else:
            ls.solve(ds.memY, ds.memU, ds.memY1)
            cc.controller(ls.x0[0], ls.x0[1], rr)
            
#       For controller time keeping            
        time.sleep(cc.interval - (time.time() - startTime))