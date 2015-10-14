#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Author: Alexander David Leech
Date:   07/10/2015
Rev:    0.3
Lang:   Python 2.7
Deps:   Pyserial, Pymodbus, scipy
"""

from pymodbus.client.sync import ModbusSerialClient as ModbusClient            #Import the MODBUS Protocol
import logging                                                                 #For debugging purposes
import time                                                                    #Time based functions
import openpyxl as xl                                                          #For data logging in Excel


#-------------------------------------------------------------------------------
# Put the below in a different file.... when I can work out how to import it!!!!

from scipy import optimize as opt
from collections import deque as dq
import numpy

class leastSquares:
    
    #Initial guess of alpha and beta    
    x0 = numpy.array([1.0,1.0])
     
    def func(self,params,y,u,Y):
        #structure of function to be minimised
        return (Y - (params[0]*y + params[1]*u))


    def solve(self,y,u,Y):
        #run least squares algorithm - add try statement here
        ans = opt.leastsq(self.func,self.x0, args=(y,u,Y))
        #translate answers
        self.x0[0] = ans[0][0]
        self.x0[1] = ans[0][1]
        #for debugging - REMOVE LATER
        print self.x0


class dataStorage:
    
    #Amount of data points used to generate alpha and beta in least squares
    dataPoints = 20
    #Initialise list to be used as 'realtime data storage'  
    memU = dq([])
    memY = dq([])
    memY1 = dq([])
    #Initial flag to signal data removal
    startupFlag = 0    

    def memoryUpdate(self,In):
        if self.memU.__len__() == 0:
            #Value for Yt+1 is not yet know so to avoid issues, set equal to zero            
            self.memU.append(In[0])
            self.memY1.append(0)            
            self.memY.append(In[3])
            print "State 1"
            
        elif (self.memU.__len__() == 1 and self.startupFlag == 0):
            #Store new data         
            self.memU.append(In[0])
            self.memY1.append(self.memY[-1])            
            self.memY.append(In[3])
            #Remove the startup data
            self.memU.popleft()
            self.memY.popleft()
            self.memY1.popleft()
            #Ensure this procedure does not run again
            self.startupFlag = 1          
            
        elif self.memU.__len__() < self.dataPoints:
            #Filling lists up to number set by dataPoints
            self.memU.append(In[0])
            self.memY1.append(self.memY[-1])            
            self.memY.append(In[3])
            
        else:
            #store data            
            self.memU.append(In[0])
            self.memY1.append(self.memY[-1])            
            self.memY.append(In[3])
            #Keep data lists at length identified
            self.memU.popleft()
            self.memY.popleft()
            self.memY1.popleft()
            
# Put the above in a different file.... when I can work out how to import it!!!!
#-------------------------------------------------------------------------------


class xlsLogging:
    
    headers = ['PV', ' ', 'SP', 'OP']
    
    def __init__(self):
        #Load workbook and create new log sheet
        self.wb = xl.load_workbook('dataLogging.xlsx')
        self.ws = self.wb.create_sheet(title='Log' + str(len(self.wb.worksheets)+1))
       
        #Setup new sheet headers
        self.ws.cell(row = 0, column = 0).value = "Date: "
        self.ws.cell(row = 0, column = 1).value = time.ctime()
        self.ws.cell(row = 2, column = 0).value = "I" 
        for x in range(0, 4):
            self.ws.cell(row= 2, column= x+1).value = self.headers[x]
		
        #Initialise Iteration Number & EPOCH time
        self.i = 1
        self.startTime = time.time()
    
    
    def writeXls(self, data):
        
        self.ws.cell(row = self.i+2, column = 0).value = (time.time() - self.startTime)
        for x in range(0, 4):
            self.ws.cell(row = self.i+2, column = x+1).value = data.getRegister(x)
		
        self.wb.save('dataLogging.xlsx')
        self.i += 1        
        return 1
                
        
class comClient:

#Note controller parameters are listed here

    interval = 5
    sp = 50
    

    comSettings = {
                    "method"   : 'rtu',
                    "port"     : 'COM3',    #COM2 for virtual serial port tests & COM3/4 for online running
                    "stopbits" : 1,                
                    "bytesize" : 8,                
                    "parity"   : 'N',
                    "baudrate" : 9600,
                    "timeout"  : 1
                  }

    def __init__(self):
        # Configure logging on the client 
        logging.basicConfig()
        self.log = logging.getLogger()
        #Change to DEBUG for full information during runtime
        self.log.setLevel(logging.INFO)
        #  Setup and Open the connection
        self.client = ModbusClient(**self.comSettings)

    def __getData(self):
        self.client.connect()
        r = self.client.read_holding_registers(0, 4, unit=0x01)               #REMEMBER: Controller is unit 0x01!!!!!
        self.client.close()
        return r

    def __writeData(self,op):
        #Set to write data to the controller output (MODBUS address 3)
        self.client.connect()
        w = self.client.write_register(3,op,unit=0x01)
        self.client.close()
        return w
           
    def controller(self,a,b,y):
        ut = ((self.sp*10)-(a*y[0]))/b
        if ut > 1000:
            self.__writeData(1000)
        elif ut < 0:
            self.__writeData(0)
        else:
            self.__writeData(ut)
        

#-------------------------------------------------------------------------------
		
if __name__ == '__main__':
    cc = comClient()
    ds = dataStorage()
    ls = leastSquares()
#    Uncomment to enable excel logging
#    datLog = xlsLogging()

    while(True):
#       For controller time keeping
        startTime = time.time()
        
#       Read controller data        
        try:
            r = cc.__getData()
        except:
		print "Modbus Error: Connection Failed"
		print "Attempting Reconnection..."

#        Uncomment to enable excel logging
#        datLog.writeXls(r)

#       Interpret data
        ds.memoryUpdate(r)
        ls.solve(ds.memY, ds.memU, ds.memY1)

#       Update Outputs
        cc.controller(ls.x0[0], ls.x0[1], r)

#       For controller time keeping            
        time.sleep(cc.interval - (time.time() - startTime))