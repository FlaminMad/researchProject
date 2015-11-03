#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Author: Alexander David Leech
Date:   12/10/2015
Rev:    1
Lang:   Python 2.7
Deps:   time, openpyxl
Desc:   Sets up data logging from the honeywell controller to an xls file.
        Ensure that the template excel file named dataLogging.xlsx is in the
        current working directory.
"""

import time                                         #Time based functions
import openpyxl as xl                               #For data logging in Excel


class xlsLogging:
    
    headers = ['PV', ' ', 'SP', 'OP','A','B']
    
    def __init__(self):
        #Load workbook and create new log sheet
        self.wb = xl.load_workbook('dataLogging.xlsx')
        self.ws = self.wb.create_sheet(title='Log' + str(len(self.wb.worksheets)+1))
       
        #Setup new sheet headers
        self.ws.cell(row = 0, column = 0).value = "Date: "
        self.ws.cell(row = 0, column = 1).value = time.ctime()
        self.ws.cell(row = 2, column = 0).value = "I" 
        for x in range(0, 6):
            self.ws.cell(row= 2, column= x+1).value = self.headers[x]
		
        #Initialise Iteration Number & EPOCH time
        self.i = 1
        self.startTime = time.time()
    
    
    def writeXls(self, data,params):
        
        self.ws.cell(row = self.i+2, column = 0).value = (time.time() - self.startTime)
        for x in range(0, 4):
            self.ws.cell(row = self.i+2, column = x+1).value = data.getRegister(x)
        self.ws.cell(row = self.i+2, column = 5).value = params[0]
        self.ws.cell(row = self.i+2, column = 6).value = params[1]
        self.wb.save('dataLogging.xlsx')
        self.i += 1        
        return 1