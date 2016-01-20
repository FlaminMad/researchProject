#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Author: Alexander David Leech
Date:   12/10/2015
Rev:    1
Lang:   Python 2.7
Deps:   time, openpyxl
Desc:   Sets up data logging from the honeywell controller to an xls file.
        Ensure that the template excel file named dataLogging.xlsx is in the
        current working directory.
"""

import os
import time                                         #Time based functions
import warnings
import openpyxl as xl                               #For data logging in Excel


class xlsLogging:
    
    headers = ['PV', ' ', 'SP', 'OP','A','B']
    
    def __init__(self,varNo):
        """
        :param varNo: vars to log (Must be 6 or 4)
        :type varNo: int
        """
        #Define OS type for compatibility
        if os.name == 'nt':
            self.osFactor = 0
        elif os.name == 'posix':
            self.osFactor = 1
        else:
            print("Unsupported OS, XLS logging may produce undesirable effects")
            self.osFactor = 1
        
        #Load workbook and create new log sheet
        try:        
            self.wb = xl.load_workbook('dataLogging.xlsx')
            print("Load Success - Creating new sheet")
            self.ws = self.wb.create_sheet(title='Log' + str(len(self.wb.worksheets)+1))
        except:
            print("Load Failed - Creating new workbook")
            self.wb = xl.workbook.Workbook()
            self.ws = self.wb.active
            self.ws.title = "Log"
        
        if varNo == 4 or varNo == 6:
            self.varNo = varNo
            #Setup some useful headers
            self.ws.cell(row = self.osFactor, column = self.osFactor).value = "Date: "
            self.ws.cell(row = self.osFactor, column = 1 + self.osFactor).value = time.ctime()
            self.ws.cell(row = 2 + self.osFactor, column = self.osFactor).value = "I" 
            for x in range(0, varNo):
                self.ws.cell(row= 2+self.osFactor, column= x+1+self.osFactor).value = self.headers[x]
        else:
            warnings.warn("Invaid number of variables - Logging Disabled")
            self.wb.save('dataLogging.xlsx')
            self.varNo = -1
               
        #Initialise Iteration Number
        self.i = 1
    
    
    def writeXls(self, startTime, data, *params):
        """
        :param startTime: Controller starting time
        :type startTime: float
        :param data: MODBUS data object to be written
        :type data: object with getRegister method
        :param *params: optional system identification array of A & B
        :type *params: list or array
        """        
        if self.varNo == -1: return             #Pass if invalid no of vars
            
        self.ws.cell(row = self.i+2+self.osFactor, column = self.osFactor).value = (time.time() - startTime)
        for x in range(0, 4):
            self.ws.cell(row = self.i+2+self.osFactor, column = x+1+self.osFactor).value = data.getRegister(x)
            
        if self.varNo == 6:
            self.ws.cell(row = self.i+2, column = 5).value = params[0][0]
            self.ws.cell(row = self.i+2, column = 6).value = params[0][1]
            
        self.wb.save('dataLogging.xlsx')
        self.i += 1     
        return