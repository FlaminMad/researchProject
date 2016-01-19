#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Author: Alexander David Leech
Date:   12/10/2015
Rev:    1
Lang:   Python 2.7
Deps:   time, openpyxl
Desc:   Sets up data logging from the honeywell controller to an xls file.
        Ensure that the template excel file named dataLogging.xlsx is in the
        current working directory.
"""

import os
import time                                         #Time based functions
import openpyxl as xl                               #For data logging in Excel


class xlsLogging4Vars:
    
    headers = ['PV', ' ', 'SP', 'OP']
    
    def __init__(self):
        #Define OS type for compatibility
        if os.name == 'nt':
            self.osFactor = 0
        elif os.name == 'posix':
            self.osFactor = 1
        else:
            print("Unsupported OS, XLS logging may produce undesirable effects")
            self.osFactor = 1
        
        #Load workbook and create new log sheet
        self.wb = xl.load_workbook('dataLogging.xlsx')
        self.ws = self.wb.create_sheet(title='Log' + str(len(self.wb.worksheets)+1))
       
        #Setup new sheet headers
        self.ws.cell(row = self.osFactor, column = self.osFactor).value = "Date: "
        self.ws.cell(row = self.osFactor, column = 1 + self.osFactor).value = time.ctime()
        self.ws.cell(row = 2 + self.osFactor, column = self.osFactor).value = "I" 
        for x in range(0, 4):
            self.ws.cell(row= 2+self.osFactor, column= x+1+self.osFactor).value = self.headers[x]
		
        #Initialise Iteration Number & EPOCH time
        self.i = 1
    
    
    def writeXls(self, data,startTime):        
        self.ws.cell(row = self.i+2+self.osFactor, column = self.osFactor).value = (time.time() - startTime)
        for x in range(0, 4):
            self.ws.cell(row = self.i+2+self.osFactor, column = x+1+self.osFactor).value = data.getRegister(x)	
        self.wb.save('dataLogging.xlsx')
        self.i += 1        
        return 1