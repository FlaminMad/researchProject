#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Author: Alexander David Leech
Date:   30/09/2015
Rev:    0.1
Lang:   Python 2.7
Deps:   Pyserial, Pymodbus
"""


from pymodbus.client.sync import ModbusSerialClient as ModbusClient            #Import the MODBUS Protocol
import logging                                                                 #For debugging purposes
import time                                                                    #For excel timestamps 
import sched                                                                   #For ensuring interval kept constant                  



class xlsLogging:
    
    headers = ['PV', ' ', 'SP', 'OP']
    
    
    def __init__(self):
        #Load workbook and create new log sheet
        import openpyxl as xl
        self.wb = xl.load_workbook('dataLogging.xlsx')
        self.ws = self.wb.create_sheet(title='Log' + str(len(self.wb.worksheets)+1))
       
        #Setup new sheet headers
        self.ws.cell(row = 0, column = 0).value = "Date: "
        self.ws.cell(row = 0, column = 1).value = time.ctime()
        self.ws.cell(row = 2, column = 0).value = "I" 
        for x in range(0, 4):
            self.ws.cell(row= 2, column= x+1).value = self.headers[x]
        
        #Initialise Iteration Number & EPOCH time
        self.i = 1
        self.startTime = time.time()
    
    
    def writeXls(self, data):
        
        self.ws.cell(row = self.i+2, column = 0).value = (time.time() - self.startTime)
        for x in range(0, 4):
            self.ws.cell(row = self.i+2, column = x+1).value = data.getRegister(x)
                    
        self.wb.save('dataLogging.xlsx')
        self.i += 1        
        return 1
        
        
        
class comClient:

    comSettings = {
                    "method"   : 'rtu',
                    "port"     : 'COM2',    #COM2 for virtual serial port tests & COM4 for online running
                    "stopbits" : 1,                
                    "bytesize" : 8,                
                    "parity"   : 'N',
                    "baudrate" : 9600,
                    "timeout"  : 1
                  }

    def __init__(self):
        # Configure logging on the client    
        logging.basicConfig()
        self.log = logging.getLogger()
        self.log.setLevel(logging.INFO)     #Change to DEBUG for full information during runtime
        #  Setup and Open the connection
        self.client = ModbusClient(**self.comSettings)

    def run(self):
        self.client.connect()
        rr = self.client.read_holding_registers(0, 4, unit=0x02)
        self.client.close()
        return rr


def sequence(cc, datLog):
    dat = cc.run()
    datLog.writeXls(dat)


if __name__ == '__main__':
    cc = comClient()
    datLog = xlsLogging()
    s = sched.scheduler(time.time, time.sleep)
    
    while(True):
        s.enter(4,1,sequence,(cc,datLog))
        s.run()
        

    
    